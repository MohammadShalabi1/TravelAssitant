import keyboard
from google import genai
import sounddevice as sd
import numpy as np
from faster_whisper import WhisperModel
from openpyxl import Workbook, load_workbook
import json, os

# ---------------- GEMINI ----------------

CLIENT_API_KEY = "////////"
client = genai.Client(api_key=CLIENT_API_KEY)

# ---------------- WHISPER MODEL ----------------
model = WhisperModel("base", device="cpu")  # or "medium" for more accuracy

# ---------------- RECORD AUDIO ----------------
samplerate = 16000

def record_audio():
    print("🎤 Recording... Hold T to speak")
    
    audio_data = []

    def callback(indata, frames, time, status):
        audio_data.append(indata.copy())

    with sd.InputStream(samplerate=samplerate, channels=1, callback=callback):
        while keyboard.is_pressed("t"):
            pass

    print("🛑 Recording stopped.")
    return np.concatenate(audio_data, axis=0)

# ---------------- TRANSCRIBE ----------------
def transcribe(audio):
    print("📝 Transcribing with Whisper...")
    audio = audio[:, 0]  # mono
    segments, _ = model.transcribe(audio, language="en")
    text = " ".join([s.text for s in segments])
    print("You said:", text)
    return text

# ---------------- ASK GEMINI ----------------
def clean_json(text):
    text = text.strip()
    if text.startswith("```"):
        text = text[text.find("{"): text.rfind("}") + 1]
    return text

def ask_gemini(text):
    prompt = f"""
    Convert the following text into JSON.

    Output must be ONLY raw JSON.
    And the Full Name should be in arabic
    JSON structure:
    {{
        "Full Name": "",
        "Comprehansion": "",
        "Dectation": "",
        "Grammer": "",
        "analysis": "",
        "Grade": "",
        "Section": ""
    }}

    Input:
    "{text}"
    """
    response = client.models.generate_content(
        model="gemini-2.5-flash", 
        contents=prompt
    )
    return clean_json(response.text)

# ---------------- WRITE TO EXCEL ----------------
def write_to_excel(json_text):
    data = json.loads(json_text)
    file = "students.xlsx"

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys()))
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active
    ws.append(list(data.values()))
    wb.save(file)

    print("✔ Added to Excel!")

# ---------------- MAIN LOOP ----------------
print("\n=== Whisper Voice to Excel ===")
print("[T] Hold to talk")
print("[I] Insert into excel")
print("[Q] Quit\n")

last_text = ""

while True:
    if keyboard.is_pressed("t"):
        audio = record_audio()
        last_text = transcribe(audio)

    if keyboard.is_pressed("i"):
        if last_text.strip() == "":
            print("⚠ Speak first!")
        else:
            print("🔄 Sending to Gemini...")
            js = ask_gemini(last_text)
            write_to_excel(js)
        while keyboard.is_pressed("i"):
            pass

    if keyboard.is_pressed("q"):
        print("👋 Exiting...")
        break
